import pandas as pd
import openpyxl
from openpyxl.styles import Alignment


def dataTo2700table(inputFile, outputFile):
    data_all = pd.read_excel(inputFile, sheet_name='输入参数')

    data_2700 = pd.DataFrame(
        columns=["边缘控制器编号", "IP地址", "主机MAC", "主机序列号", "板卡编号", "板卡出厂编号", "板卡类型",
                 "板卡是否启用",
                 "通道编号", "测点（通道）类型", "设备名称", "测点（点位）名称", "键相类型", "工作转速", "电机额定转速",
                 "电机同步转速", "电源频率", "电机转子条数", "轴承型号", "轴承生产厂家", "齿轮齿数Z", "叶轮叶片数目",
                 "导叶叶片数目"]
    )

    rows = []
    for index, row in data_all.iterrows():
        if row["通道编码"][:6] == "50294D":
            row_data = {"边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                        "主机MAC": row["通道编码"][:-3], "板卡编号": "C" + row["通道编码"][-3:-1],
                        "通道编号": "CH0" + row["通道编码"][-1], "测点（通道）类型": row["传感器类型"],
                        "设备名称": row["设备名称"], "测点（点位）名称": row["测点名称"], "工作转速": row["工作转速"],
                        "电机额定转速": row["电机额定转速"], "电机同步转速": row["电机同步转速"],
                        "电源频率": row["电源频率"], "电机转子条数": row["电机转子条数"], "轴承型号": row["轴承型号"],
                        "轴承生产厂家": row["轴承生产厂家"], "齿轮齿数Z": row["齿轮齿数Z"],
                        "叶轮叶片数目": row["叶轮叶片数目"], "导叶叶片数目": row["导叶叶片数目"],
                        "板卡类型": "高速卡" if row["传感器类型"] in ["加速度", "速度", "位移", "转速"] else "低速卡",
                        "板卡是否启用": "是" if row["测点名称"] else "否",
                        "键相类型": "虚拟键相" if row["工作转速"] and row["工作转速"] != "/" else None}

            rows.append(row_data)

    data_2700 = pd.DataFrame(rows, columns=data_2700.columns)
    data_2700 = data_2700.sort_values(by=["主机MAC", "板卡编号", "通道编号"])

    with pd.ExcelWriter(outputFile, engine='openpyxl') as writer:
        for mac, group in data_2700.groupby('主机MAC'):
            group.to_excel(writer, index=False, sheet_name=f"{mac}")

    workbook = openpyxl.load_workbook(outputFile)

    for sheet in workbook.worksheets:
        if sheet.title.startswith("50294D"):
            # 合并第前四列中相同的数值的单元格
            merge_cells_in_column_1(sheet, column_index=3)
            # 合并第五-八列中相同的数值的单元格
            merge_cells_in_column_2(sheet, column_index=5)

    workbook.save(outputFile)


def merge_cells_in_column_1(sheet, column_index):
    start_row = 2  # 从第2行开始（第1行是表头）
    end_row = sheet.max_row

    current_value = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value != current_value:
            if merge_start and row - 1 > merge_start:
                sheet.merge_cells(start_row=merge_start, start_column=column_index+1,
                                  end_row=row - 1, end_column=column_index+1)
                align_merged_cells(sheet, merge_start, row - 1, column_index+1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index,
                                  end_row=row - 1, end_column=column_index)
                align_merged_cells(sheet, merge_start, row - 1, column_index)
                sheet.merge_cells(start_row=merge_start, start_column=column_index-1,
                                  end_row=row - 1, end_column=column_index-1)
                align_merged_cells(sheet, merge_start, row - 1, column_index-1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index-2,
                                  end_row=row - 1, end_column=column_index-2)
                align_merged_cells(sheet, merge_start, row - 1, column_index-2)
            current_value = cell.value
            merge_start = row
        elif row == end_row:  # 如果到达最后一行
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                              end_row=row, end_column=column_index + 1)
            align_merged_cells(sheet, merge_start, row, column_index + 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index,
                              end_row=row, end_column=column_index)
            align_merged_cells(sheet, merge_start, row, column_index)
            sheet.merge_cells(start_row=merge_start, start_column=column_index - 1,
                              end_row=row, end_column=column_index - 1)
            align_merged_cells(sheet, merge_start, row, column_index - 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index - 2,
                              end_row=row, end_column=column_index - 2)
            align_merged_cells(sheet, merge_start, row, column_index - 2)


def merge_cells_in_column_2(sheet, column_index):
    start_row = 2  # 从第2行开始（第1行是表头）
    end_row = sheet.max_row

    current_value = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value != current_value:
            if merge_start and row - 1 > merge_start:
                sheet.merge_cells(start_row=merge_start, start_column=column_index,
                                  end_row=row - 1, end_column=column_index)
                align_merged_cells(sheet, merge_start, row - 1, column_index)
                sheet.merge_cells(start_row=merge_start, start_column=column_index+1,
                                  end_row=row - 1, end_column=column_index+1)
                align_merged_cells(sheet, merge_start, row - 1, column_index+1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index+2,
                                  end_row=row - 1, end_column=column_index+2)
                align_merged_cells(sheet, merge_start, row - 1, column_index+2)
                sheet.merge_cells(start_row=merge_start, start_column=column_index+3,
                                  end_row=row - 1, end_column=column_index+3)
                align_merged_cells(sheet, merge_start, row - 1, column_index+3)
            current_value = cell.value
            merge_start = row
        elif row == end_row:  # 如果到达最后一行
            sheet.merge_cells(start_row=merge_start, start_column=column_index,
                              end_row=row, end_column=column_index)
            align_merged_cells(sheet, merge_start, row, column_index)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                              end_row=row, end_column=column_index + 1)
            align_merged_cells(sheet, merge_start, row, column_index + 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 2,
                              end_row=row, end_column=column_index + 2)
            align_merged_cells(sheet, merge_start, row, column_index + 2)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 3,
                              end_row=row, end_column=column_index + 3)
            align_merged_cells(sheet, merge_start, row, column_index + 3)


def align_merged_cells(sheet, start_row, end_row, column_index):
    """对齐合并后的单元格"""
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        cell.alignment = Alignment(horizontal='center', vertical='center')


if __name__ == "__main__":
    dataTo2700table("test/excel/data_all.xlsx", "test/excel/data_2700.xlsx")
